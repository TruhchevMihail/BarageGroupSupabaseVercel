import io
import json
from pathlib import Path

import app as app_module
from openpyxl import load_workbook
from werkzeug.datastructures import FileStorage

from barage_app.services.assets_csv import (
    build_asset_csv_template,
    build_asset_xlsx_template,
    export_assets_csv,
    export_assets_xlsx,
    parse_asset_csv_upload,
    parse_asset_import_upload,
)


def _asset(inventory_number, *, name='Машина', brand='Brand', model='Model', location=None, status=None):
    return app_module.Asset(
        inventory_number=inventory_number,
        name=name,
        brand=brand,
        model=model,
        current_location_id=location.id if location else None,
        status=status or app_module.STATUS_WAREHOUSE,
    )


def test_authenticated_user_can_export_assets_csv_with_filters(client, db, make_user, login):
    warehouse = app_module.Location(name='Склад CSV', type=app_module.LOC_WAREHOUSE, is_active=True)
    service = app_module.Location(name='Сервиз CSV', type=app_module.LOC_SERVICE, is_active=True)
    db.session.add_all([warehouse, service])
    db.session.commit()
    db.session.add_all([
        _asset('CSV-1', name='Къртач', location=warehouse, status=app_module.STATUS_WAREHOUSE),
        _asset('CSV-2', name='Дрелка', location=service, status=app_module.STATUS_SERVICE),
    ])
    db.session.commit()

    user = make_user(full_name='CSV User', email='csv-user@example.com', role=app_module.ROLE_USER)
    login(user)

    response = client.get('/assets/export.csv?status=В сервиз')
    payload = response.get_data()
    body = payload.decode('utf-8-sig')

    assert response.status_code == 200
    assert response.mimetype == 'text/csv'
    assert payload.startswith(b'\xef\xbb\xbf')
    assert body.splitlines()[0] == 'sep=;'
    assert response.headers['Content-Type'] == 'text/csv; charset=utf-8'
    assert 'attachment; filename="assets_export.csv"' in response.headers['Content-Disposition']
    assert 'Инвентарен №' in body
    assert 'Сервиз CSV' in body
    assert 'CSV-2' in body
    assert 'CSV-1' not in body
    assert 'CSV-2;Дрелка;' in body


def test_authenticated_user_can_export_assets_xlsx_with_filters(client, db, make_user, login):
    warehouse = app_module.Location(name='Склад XLSX', type=app_module.LOC_WAREHOUSE, is_active=True)
    service = app_module.Location(name='Сервиз XLSX', type=app_module.LOC_SERVICE, is_active=True)
    db.session.add_all([warehouse, service])
    db.session.commit()
    db.session.add_all([
        _asset('XLSX-1', name='Къртач', location=warehouse, status=app_module.STATUS_WAREHOUSE),
        _asset('XLSX-2', name='Дрелка', location=service, status=app_module.STATUS_SERVICE),
    ])
    db.session.commit()

    user = make_user(full_name='Excel User', email='excel-user@example.com', role=app_module.ROLE_USER)
    login(user)

    response = client.get('/assets/export.xlsx?status=В сервиз')
    workbook = load_workbook(io.BytesIO(response.get_data()))
    sheet = workbook.active

    assert response.status_code == 200
    assert response.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    assert 'attachment; filename="assets_export.xlsx"' in response.headers['Content-Disposition']
    assert sheet.title == 'Машини'
    assert sheet.freeze_panes == 'A2'
    assert sheet.auto_filter.ref == sheet.dimensions
    assert [cell.value for cell in sheet[1]][:4] == ['Инвентарен №', 'Име', 'Марка', 'Модел']
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    assert rows[0][0] == 'XLSX-2'
    assert rows[0][1] == 'Дрелка'
    assert rows[0][2] == 'Brand'
    assert rows[0][3] == 'Model'
    assert rows[0][5] == 'Машина'
    assert rows[0][7] == 'Сервиз XLSX'
    assert rows[0][8] == 'Сервиз'
    assert rows[0][9] == 'В сервиз'


def test_unauthenticated_user_cannot_export_assets_csv(client):
    response = client.get('/assets/export.csv')
    assert response.status_code == 302
    assert '/login' in response.headers['Location']


def test_unauthenticated_user_cannot_export_assets_xlsx(client):
    response = client.get('/assets/export.xlsx')
    assert response.status_code == 302
    assert '/login' in response.headers['Location']


def test_assets_page_shows_import_only_to_admin(client, db, make_user, login):
    user = make_user(full_name='Regular User', email='regular-assets@example.com', role=app_module.ROLE_USER)
    admin = make_user(full_name='Admin User', email='admin-assets@example.com', role=app_module.ROLE_SUPERUSER)

    login(user)
    user_page = client.get('/assets')
    assert user_page.status_code == 200
    assert 'Експорт Excel' in user_page.get_data(as_text=True)
    assert 'Експорт CSV' in user_page.get_data(as_text=True)
    assert 'Импорт CSV/Excel' not in user_page.get_data(as_text=True)

    login(admin)
    admin_page = client.get('/assets')
    assert admin_page.status_code == 200
    assert 'Импорт CSV/Excel' in admin_page.get_data(as_text=True)


def test_non_admin_cannot_access_asset_import(client, db, make_user, login):
    user = make_user(full_name='Import User', email='import-user@example.com', role=app_module.ROLE_USER)
    login(user)
    response = client.get('/assets/import')
    assert response.status_code == 302
    assert '/dashboard' in response.headers['Location']


def test_admin_can_access_asset_import_page(client, db, make_user, login):
    admin = make_user(full_name='Import Admin', email='import-admin@example.com', role=app_module.ROLE_SUPERUSER)
    login(admin)
    response = client.get('/assets/import')
    html = response.get_data(as_text=True)
    assert response.status_code == 200
    assert 'Качи CSV или Excel файл' in html
    assert 'Импортът е достъпен само за администратори.' in html
    assert 'Можете да качите CSV или Excel файл.' in html


def test_asset_import_preview_validates_rows_and_does_not_touch_uploads(client, app, db, make_user, login, default_csrf):
    admin = make_user(full_name='Preview Admin', email='preview-admin@example.com', role=app_module.ROLE_SUPERUSER)
    location = app_module.Location(name='Централен склад', type=app_module.LOC_WAREHOUSE, is_active=True)
    db.session.add(location)
    db.session.commit()
    login(admin)

    before_uploads = set(Path(app.config['UPLOAD_FOLDER']).rglob('*'))
    csv_text = (
        '\ufeffsep=;\n'
        'Инвентарен №;Име;Марка;Модел;Вид актив;Текуща локация\n'
        'IMP-1;Къртач;Bosch;GSH 11;Машина;Централен склад\n'
        'IMP-1;Дубликат;Bosch;GSH 12;Машина;Централен склад\n'
        'IMP-2;Без локация;Bosch;GSH 13;Невалиден;Несъществуваща локация\n'
    )
    response = client.post(
        '/assets/import/preview',
        data={
            'csrf_token': default_csrf,
            'csv_file': (io.BytesIO(csv_text.encode('utf-8')), 'assets.csv'),
        },
        content_type='multipart/form-data',
    )
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert 'Валидни редове' in html
    assert '<strong>1</strong>' in html
    assert 'Дублиран инвентарен № в импорт файла.' in html
    assert 'Невалиден вид актив.' in html
    assert app_module.Asset.query.count() == 0
    assert set(Path(app.config['UPLOAD_FOLDER']).rglob('*')) == before_uploads


def test_asset_xlsx_import_preview_validates_rows_and_does_not_touch_uploads(client, app, db, make_user, login, default_csrf):
    admin = make_user(full_name='Excel Preview Admin', email='excel-preview-admin@example.com', role=app_module.ROLE_SUPERUSER)
    location = app_module.Location(name='Склад Excel', type=app_module.LOC_WAREHOUSE, is_active=True)
    db.session.add(location)
    db.session.commit()
    login(admin)

    workbook = load_workbook(io.BytesIO(build_asset_xlsx_template()))
    sheet = workbook.active
    sheet.append(['XL-1', 'Къртач', 'Bosch', 'GSH 11', 'Къртач', 'Машина', 'SN1', 'Склад Excel'])
    sheet.append(['XL-1', 'Дубликат', 'Bosch', 'GSH 12', 'Къртач', 'Машина', 'SN2', 'Склад Excel'])
    sheet.append(['XL-2', '', 'Bosch', 'GSH 13', 'Къртач', 'Невалиден', 'SN3', 'Несъществуваща локация'])
    payload = io.BytesIO()
    workbook.save(payload)
    payload.seek(0)

    before_uploads = set(Path(app.config['UPLOAD_FOLDER']).rglob('*'))
    response = client.post(
        '/assets/import/preview',
        data={
            'csrf_token': default_csrf,
            'csv_file': (payload, 'assets.xlsx'),
        },
        content_type='multipart/form-data',
    )
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert 'Валидни редове' in html
    assert '<strong>1</strong>' in html
    assert 'Дублиран инвентарен № в импорт файла.' in html
    assert 'Липсва име/тип на машината.' in html
    assert 'Невалиден вид актив.' in html
    assert app_module.Asset.query.count() == 0
    assert set(Path(app.config['UPLOAD_FOLDER']).rglob('*')) == before_uploads


def test_asset_import_rejects_non_csv_upload(client, db, make_user, login, default_csrf):
    admin = make_user(full_name='Reject Admin', email='reject-admin@example.com', role=app_module.ROLE_SUPERUSER)
    login(admin)

    response = client.post(
        '/assets/import/preview',
        data={
            'csrf_token': default_csrf,
            'csv_file': (io.BytesIO(b'not csv'), 'assets.txt'),
        },
        content_type='multipart/form-data',
    )

    assert response.status_code == 200
    assert 'Файлът трябва да бъде CSV или Excel (.xlsx).' in response.get_data(as_text=True)
    assert app_module.Asset.query.count() == 0


def test_export_service_returns_excel_friendly_bytes_with_cyrillic_and_semicolons(db):
    warehouse = app_module.Location(name='База; Централна', type=app_module.LOC_WAREHOUSE, is_active=True)
    db.session.add(warehouse)
    db.session.commit()
    db.session.add(_asset('CSV-EX-1', name='Машина "Тест"', brand='Марка', model='Модел', location=warehouse))
    db.session.commit()

    payload = export_assets_csv({'sort': 'inventory', 'direction': 'asc'})
    text = payload.decode('utf-8-sig')
    lines = text.splitlines()

    assert payload.startswith(b'\xef\xbb\xbf')
    assert lines[0] == 'sep=;'
    assert lines[1].startswith('Инвентарен №;Име;Марка;Модел;')
    assert 'CSV-EX-1;"Машина ""Тест""";Марка;Модел;' in text
    assert '"База; Централна"' in text


def test_export_service_returns_excel_workbook_with_cyrillic_and_filters(db):
    warehouse = app_module.Location(name='База Excel', type=app_module.LOC_WAREHOUSE, is_active=True)
    service = app_module.Location(name='Сервиз Excel', type=app_module.LOC_SERVICE, is_active=True)
    db.session.add_all([warehouse, service])
    db.session.commit()
    db.session.add_all([
        _asset('EXCEL-1', name='Машина', brand='Марка', model='Модел', location=warehouse, status=app_module.STATUS_WAREHOUSE),
        _asset('EXCEL-2', name='Сервизна', brand='Марка', model='Модел', location=service, status=app_module.STATUS_SERVICE),
    ])
    db.session.commit()

    payload = export_assets_xlsx({'status': 'В сервиз', 'sort': 'inventory', 'direction': 'asc'})
    workbook = load_workbook(io.BytesIO(payload))
    sheet = workbook.active

    assert sheet.title == 'Машини'
    assert sheet['A1'].font.bold is True
    assert sheet.freeze_panes == 'A2'
    assert [cell.value for cell in sheet[1]] == [
        'Инвентарен №',
        'Име',
        'Марка',
        'Модел',
        'Категория',
        'Вид актив',
        'Сериен №',
        'Текуща локация',
        'Тип локация',
        'Статус',
        'Дни в сервиз',
        'Последно преместване',
        'Дата на създаване',
    ]
    values = list(sheet.iter_rows(min_row=2, values_only=True))
    assert len(values) == 1
    assert values[0][0] == 'EXCEL-2'
    assert values[0][1] == 'Сервизна'
    assert values[0][7] == 'Сервиз Excel'


def test_import_parser_accepts_exported_excel_friendly_csv(db):
    location = app_module.Location(name='Централен склад', type=app_module.LOC_WAREHOUSE, is_active=True)
    db.session.add(location)
    db.session.commit()

    payload = build_asset_csv_template().decode('utf-8-sig').replace('INV-001', 'IMPORT-1')
    upload = FileStorage(stream=io.BytesIO(payload.encode('utf-8-sig')), filename='assets.csv')

    preview = parse_asset_csv_upload(upload)

    assert preview.errors == []
    assert preview.error_rows == []
    assert preview.valid_count == 1
    assert preview.valid_rows[0].data['inventory_number'] == 'IMPORT-1'


def test_import_parser_accepts_exported_xlsx_template(db):
    location = app_module.Location(name='Централен склад', type=app_module.LOC_WAREHOUSE, is_active=True)
    db.session.add(location)
    db.session.commit()

    workbook = load_workbook(io.BytesIO(build_asset_xlsx_template()))
    sheet = workbook.active
    sheet.append(['IMPORT-XLSX-1', 'Къртач', 'Bosch', 'GSH 11', 'Къртач', 'Машина', 'SN123', 'Централен склад'])
    payload = io.BytesIO()
    workbook.save(payload)
    payload.seek(0)
    upload = FileStorage(stream=payload, filename='assets.xlsx')

    preview = parse_asset_import_upload(upload)

    assert preview.errors == []
    assert preview.error_rows == []
    assert preview.valid_count == 1
    assert preview.valid_rows[0].data['inventory_number'] == 'IMPORT-XLSX-1'


def test_import_parser_accepts_comma_delimited_utf8_csv(db):
    location = app_module.Location(name='Склад Запад', type=app_module.LOC_WAREHOUSE, is_active=True)
    db.session.add(location)
    db.session.commit()
    csv_text = (
        'inventory_number,name,brand,model,asset_type,current_location\n'
        'CSV-COMMA-1,Ъглошлайф,Bosch,GWS,Машина,Склад Запад\n'
    )
    upload = FileStorage(stream=io.BytesIO(csv_text.encode('utf-8')), filename='assets.csv')

    preview = parse_asset_csv_upload(upload)

    assert preview.errors == []
    assert preview.error_rows == []
    assert preview.valid_rows[0].data['inventory_number'] == 'CSV-COMMA-1'


def test_import_parser_accepts_cp1251_csv(db):
    location = app_module.Location(name='Склад Изток', type=app_module.LOC_WAREHOUSE, is_active=True)
    db.session.add(location)
    db.session.commit()
    csv_text = (
        'Инвентарен №;Име;Марка;Модел;Вид актив;Обект\n'
        'CSV-CP1251-1;Къртач;Bosch;GSH 5;Машина;Склад Изток\n'
    )
    upload = FileStorage(stream=io.BytesIO(csv_text.encode('cp1251')), filename='assets.csv')

    preview = parse_asset_csv_upload(upload)

    assert preview.errors == []
    assert preview.error_rows == []
    assert preview.valid_rows[0].data['current_location_name'] == 'Склад Изток'


def test_import_parser_rejects_unknown_encoding_with_clear_message():
    upload = FileStorage(stream=io.BytesIO(b'\x98\x99\x9a\x9b'), filename='assets.csv')

    preview = parse_asset_csv_upload(upload)

    assert preview.valid_rows == []
    assert preview.error_rows == []
    assert preview.errors == ['Файлът не може да бъде прочетен. Запазете го като CSV UTF-8 и опитайте отново.']


def test_import_parser_requires_inventory_column(db):
    location = app_module.Location(name='Склад Юг', type=app_module.LOC_WAREHOUSE, is_active=True)
    db.session.add(location)
    db.session.commit()
    csv_text = (
        'Име;Марка;Текуща локация\n'
        'Машина;Bosch;Склад Юг\n'
    )
    upload = FileStorage(stream=io.BytesIO(csv_text.encode('utf-8-sig')), filename='assets.csv')

    preview = parse_asset_csv_upload(upload)

    assert preview.errors == ['Липсва задължителна колона: Инвентарен №']


def test_admin_can_download_asset_xlsx_template(client, db, make_user, login):
    admin = make_user(full_name='Template Admin', email='template-admin@example.com', role=app_module.ROLE_SUPERUSER)
    login(admin)

    response = client.get('/assets/import/template.xlsx')
    workbook = load_workbook(io.BytesIO(response.get_data()))
    sheet = workbook.active

    assert response.status_code == 200
    assert response.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    assert 'attachment; filename="assets_import_template.xlsx"' in response.headers['Content-Disposition']
    assert sheet.title == 'Машини'
    assert sheet.freeze_panes == 'A2'
    assert [cell.value for cell in sheet[1]] == ['Инвентарен №', 'Име', 'Марка', 'Модел', 'Категория', 'Вид актив', 'Сериен №', 'Текуща локация']
    assert sheet.max_row == 1


def test_asset_import_confirm_upserts_valid_rows_without_deleting_assets(client, db, make_user, login, default_csrf):
    admin = make_user(full_name='Confirm Admin', email='confirm-admin@example.com', role=app_module.ROLE_SUPERUSER)
    location = app_module.Location(name='Склад Импорт', type=app_module.LOC_WAREHOUSE, is_active=True)
    db.session.add(location)
    db.session.commit()
    existing = _asset('KEEP-1', name='Стара машина', brand='Old', model='Old', location=location)
    untouched = _asset('KEEP-2', name='Остава', brand='Same', model='Same', location=location)
    db.session.add_all([existing, untouched])
    db.session.commit()
    login(admin)

    payload = json.dumps([
        {
            'inventory_number': 'KEEP-1',
            'name': 'Обновена машина',
            'brand': 'New',
            'model': 'New',
            'asset_type': 'Машина',
            'current_location': 'Склад Импорт',
        },
        {
            'inventory_number': 'NEW-1',
            'name': 'Нова машина',
            'brand': 'Brand',
            'model': 'Model',
            'asset_type': 'Инструмент',
            'current_location': 'Склад Импорт',
        },
    ], ensure_ascii=False)

    response = client.post(
        '/assets/import/confirm',
        data={'csrf_token': default_csrf, 'preview_payload': payload},
        follow_redirects=False,
    )

    assert response.status_code == 302
    assert app_module.Asset.query.count() == 3
    db.session.refresh(existing)
    db.session.refresh(untouched)
    created = app_module.Asset.query.filter_by(inventory_number='NEW-1').one()
    assert existing.name == 'Обновена машина'
    assert untouched.name == 'Остава'
    assert created.asset_type == 'Инструмент'
