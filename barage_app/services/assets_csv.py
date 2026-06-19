import csv
import io
import json
from dataclasses import dataclass, field
from datetime import date, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import Integer, cast, or_
from sqlalchemy.orm import joinedload

from barage_app.constants import (
    ASSET_TYPE_OPTIONS,
    LOCATION_META,
    LOCATION_TYPE_TO_STATUS,
    STATUS_TO_LOCATION_TYPE,
    STATUS_WAREHOUSE,
)
from barage_app.extensions import db
from barage_app.models import Asset, AssetHistory, Location, User
from barage_app.services.service_stay import apply_long_service_stay_filter, enrich_assets_with_service_stay


ASSET_CSV_MAX_BYTES = 1024 * 1024
ASSET_XLSX_MAX_BYTES = 5 * 1024 * 1024
ASSET_CSV_MAX_ROWS = 1000
ASSET_CSV_ENCODING = 'utf-8-sig'
ASSET_CSV_DELIMITER = ';'
ASSET_CSV_SEPARATOR_HINT = f'sep={ASSET_CSV_DELIMITER}'
ASSET_CSV_REQUIRED_COLUMN_LABEL = 'Инвентарен №'
ASSET_EXPORT_HEADERS = [
    'Инвентарен №',
    'Име',
    'Марка',
    'Модел',
    'Категория',
    'Вид актив',
    'Сериен №',
    'Текуща локация',
    'Тип локация',
    'Статус',
    'Дни в сервиз',
    'Последно преместване',
    'Дата на създаване',
]
ASSET_CSV_TEMPLATE_HEADERS = ['Инвентарен №', 'Име', 'Марка', 'Модел', 'Категория', 'Вид актив', 'Сериен №', 'Текуща локация']
ASSET_XLSX_SHEET_NAME = 'Машини'

HEADER_ALIASES = {
    '№': 'inventory_number',
    'no': 'inventory_number',
    'number': 'inventory_number',
    'inventory': 'inventory_number',
    'inventory_number': 'inventory_number',
    'инвентарен №': 'inventory_number',
    'инвентарен номер': 'inventory_number',
    'име': 'name',
    'тип / име': 'name',
    'тип': 'name',
    'name': 'name',
    'brand': 'brand',
    'марка': 'brand',
    'model': 'model',
    'модел': 'model',
    'category': 'category',
    'категория': 'category',
    'asset_type': 'asset_type',
    'вид актив': 'asset_type',
    'serial_number': 'serial_number',
    'сериен №': 'serial_number',
    'сериен номер': 'serial_number',
    'current_location': 'current_location',
    'current_location_id': 'current_location',
    'локация': 'current_location',
    'обект': 'current_location',
    'текуща локация': 'current_location',
    'purchase_date': 'purchase_date',
    'дата на закупуване': 'purchase_date',
}


@dataclass
class AssetCsvRow:
    row_number: int
    action: str
    data: dict
    errors: list[str] = field(default_factory=list)


@dataclass
class AssetCsvPreview:
    valid_rows: list[AssetCsvRow] = field(default_factory=list)
    error_rows: list[AssetCsvRow] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    @property
    def valid_count(self):
        return len(self.valid_rows)

    @property
    def error_count(self):
        return len(self.error_rows)

    def payload(self):
        return json.dumps([row.data for row in self.valid_rows], ensure_ascii=False)


def _normalize_header(value):
    return HEADER_ALIASES.get((value or '').strip().lower(), (value or '').strip())


def _clean(value):
    return str(value or '').strip()


def _build_csv_text(header_row, data_rows):
    output = io.StringIO(newline='')
    output.write(f'{ASSET_CSV_SEPARATOR_HINT}\r\n')
    writer = csv.writer(output, delimiter=ASSET_CSV_DELIMITER, quoting=csv.QUOTE_MINIMAL, lineterminator='\r\n')
    writer.writerow(header_row)
    writer.writerows(data_rows)
    return output.getvalue()


def _build_csv_bytes(header_row, data_rows):
    return _build_csv_text(header_row, data_rows).encode(ASSET_CSV_ENCODING)


def _autosize_worksheet(worksheet):
    for column_cells in worksheet.columns:
        values = [len(str(cell.value)) for cell in column_cells if cell.value not in (None, '')]
        max_length = max(values, default=0)
        letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 32)


def _build_xlsx_bytes(sheet_name, header_row, data_rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(header_row)
    for row in data_rows:
        worksheet.append(row)

    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.font = header_font

    worksheet.freeze_panes = 'A2'
    worksheet.auto_filter.ref = worksheet.dimensions
    _autosize_worksheet(worksheet)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def _decode_csv_bytes(raw_bytes):
    for encoding in ('utf-8-sig', 'utf-8', 'cp1251'):
        try:
            return raw_bytes.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    return None, None


def _detect_csv_delimiter(lines):
    if not lines:
        return None, 'CSV файлът няма заглавен ред.'

    first_line = lines[0].strip().lstrip('\ufeff')
    data_lines = lines
    if first_line.lower() == ASSET_CSV_SEPARATOR_HINT:
        data_lines = lines[1:]
        if not data_lines:
            return None, 'CSV файлът няма заглавен ред.'
        return ASSET_CSV_DELIMITER, None

    sample = '\n'.join(line for line in data_lines[:5] if line.strip())
    if not sample:
        return None, 'CSV файлът няма заглавен ред.'

    semicolons = sample.count(';')
    commas = sample.count(',')
    if semicolons and not commas:
        return ';', None
    if commas and not semicolons:
        return ',', None
    if semicolons and commas:
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=';,')
        except csv.Error:
            return None, 'Неуспешно разпознаване на разделителя. Използвайте ";" или ",".'
        return dialect.delimiter, None
    return None, 'Неуспешно разпознаване на разделителя. Използвайте ";" или ",".'


def _prepare_csv_reader(text):
    lines = text.splitlines()
    delimiter, delimiter_error = _detect_csv_delimiter(lines)
    if delimiter_error:
        return None, delimiter_error

    start_index = 1 if lines and lines[0].strip().lstrip('\ufeff').lower() == ASSET_CSV_SEPARATOR_HINT else 0
    csv_text = '\n'.join(lines[start_index:])
    reader = csv.DictReader(io.StringIO(csv_text, newline=''), delimiter=delimiter)
    if not reader.fieldnames:
        return None, 'CSV файлът няма заглавен ред.'
    reader.fieldnames = [_normalize_header(header) for header in reader.fieldnames]
    return reader, None


def _normalize_excel_cell(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    return str(value).strip()


def _prepare_xlsx_rows(raw_bytes):
    try:
        workbook = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception:
        return None, 'Файлът не може да бъде прочетен. Запазете го като Excel (.xlsx) и опитайте отново.'

    worksheet = workbook.worksheets[0] if workbook.worksheets else None
    if worksheet is None:
        return None, 'Excel файлът няма работен лист.'

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return None, 'Excel файлът няма заглавен ред.'

    header_row = [_normalize_header(_normalize_excel_cell(cell)) for cell in rows[0]]
    if not any(header_row):
        return None, 'Excel файлът няма заглавен ред.'

    data_rows = []
    for source_row in rows[1:]:
        row = {}
        has_value = False
        for index, header in enumerate(header_row):
            if not header:
                continue
            value = _normalize_excel_cell(source_row[index] if index < len(source_row) else '')
            if value:
                has_value = True
            row[header] = value
        if has_value:
            data_rows.append(row)
    return {'fieldnames': header_row, 'rows': data_rows}, None


def _parse_date(value):
    value = _clean(value)
    if not value:
        return None, None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date(), None
    except ValueError:
        return None, 'Невалидна дата. Използвайте формат YYYY-MM-DD.'


def normalize_asset_status_filter(value):
    value = (value or '').strip()
    return LOCATION_TYPE_TO_STATUS.get(STATUS_TO_LOCATION_TYPE.get(value, ''), value)


def asset_status_from_location(location):
    if not location:
        return STATUS_WAREHOUSE
    return LOCATION_TYPE_TO_STATUS.get(location.type, STATUS_WAREHOUSE)


def asset_display_status(asset):
    if asset.current_location:
        return asset_status_from_location(asset.current_location)
    return asset.status


def parse_asset_filter_args(args):
    direction = args.get('direction', 'asc').lower().strip()
    if direction not in ('asc', 'desc'):
        direction = 'asc'
    return {
        'q': args.get('q', '').strip(),
        'status': normalize_asset_status_filter(args.get('status', '').strip()),
        'location_id': args.get('location', type=int) or args.get('location_id', type=int),
        'category': args.get('category', '').strip(),
        'asset_type': args.get('asset_type', '').strip(),
        'condition': args.get('condition', '').strip(),
        'responsible_user_id': args.get('responsible_user_id', type=int),
        'service_stay': args.get('service_stay', '').strip(),
        'sort': args.get('sort', 'inventory').strip(),
        'direction': direction,
    }


def build_assets_query(filters):
    query = Asset.query.options(
        joinedload(Asset.current_location),
        joinedload(Asset.created_by),
        joinedload(Asset.responsible_user),
    )
    q = filters.get('q')
    if q:
        like = f'%{q}%'
        query = query.filter(or_(
            Asset.inventory_number.ilike(like),
            Asset.name.ilike(like),
            Asset.alias_name.ilike(like),
            Asset.brand.ilike(like),
            Asset.model.ilike(like),
            Asset.serial_number.ilike(like),
            Asset.company_name.ilike(like),
            Asset.supplier_company.ilike(like),
            Asset.invoice_number.ilike(like),
            Asset.notes.ilike(like),
            Asset.created_by.has(User.full_name.ilike(like)),
            Asset.responsible_user.has(User.full_name.ilike(like)),
            Asset.current_location.has(or_(
                Location.name.ilike(like),
                Location.city.ilike(like),
                Location.address.ilike(like),
            )),
        ))

    status_location_type = STATUS_TO_LOCATION_TYPE.get(filters.get('status'))
    if status_location_type:
        query = query.join(Location, Asset.current_location_id == Location.id).filter(Location.type == status_location_type)
    if filters.get('location_id'):
        query = query.filter_by(current_location_id=filters['location_id'])
    if filters.get('category'):
        query = query.filter(Asset.category == filters['category'])
    if filters.get('asset_type'):
        query = query.filter(Asset.asset_type == filters['asset_type'])
    if filters.get('condition'):
        query = query.filter_by(condition=filters['condition'])
    if filters.get('responsible_user_id'):
        query = query.filter_by(responsible_user_id=filters['responsible_user_id'])
    if filters.get('service_stay') == 'long':
        query = apply_long_service_stay_filter(query)
    return query


def order_assets_query(query, sort, direction):
    inventory_order = [cast(Asset.inventory_number, Integer), Asset.inventory_number]
    sort_map = {
        'inventory': inventory_order,
        'type': [Asset.asset_type],
        'brand': [Asset.brand],
        'model': [Asset.model],
        'serial': [Asset.serial_number],
        'purchase_date': [Asset.purchase_date],
        'created_at': [Asset.created_at],
    }
    columns = sort_map.get(sort, inventory_order)
    order_by = tuple(col.desc() if direction == 'desc' else col.asc() for col in columns)
    return query.order_by(*order_by)


def format_datetime(value):
    return value.strftime('%Y-%m-%d %H:%M') if value else ''


def format_date(value):
    return value.strftime('%Y-%m-%d') if value else ''


def _asset_export_rows(assets):
    rows = []
    for asset in assets:
        location = asset.current_location
        location_label = LOCATION_META.get(location.type, {}).get('label', location.type) if location else ''
        service_days = asset.service_stay_days if getattr(asset, 'service_stay_is_long', False) else ''
        rows.append([
            asset.inventory_number,
            asset.name,
            asset.brand,
            asset.model,
            asset.category or '',
            asset.asset_type or '',
            asset.serial_number or '',
            location.name if location else '',
            location_label,
            asset_display_status(asset),
            service_days,
            format_datetime(asset.last_moved_at),
            format_datetime(asset.created_at),
        ])
    return rows


def export_assets_csv(filters):
    query = order_assets_query(build_assets_query(filters), filters.get('sort'), filters.get('direction'))
    assets = query.all()
    enrich_assets_with_service_stay(assets)
    return _build_csv_bytes(ASSET_EXPORT_HEADERS, _asset_export_rows(assets))


def export_assets_xlsx(filters):
    query = order_assets_query(build_assets_query(filters), filters.get('sort'), filters.get('direction'))
    assets = query.all()
    enrich_assets_with_service_stay(assets)
    return _build_xlsx_bytes(ASSET_XLSX_SHEET_NAME, ASSET_EXPORT_HEADERS, _asset_export_rows(assets))


def build_asset_csv_template():
    return _build_csv_bytes(
        ASSET_CSV_TEMPLATE_HEADERS,
        [['INV-001', 'Къртач', 'Bosch', 'GSH 11', 'Къртач', 'Машина', 'SN123', 'Централен склад']],
    )


def build_asset_xlsx_template():
    return _build_xlsx_bytes(ASSET_XLSX_SHEET_NAME, ASSET_CSV_TEMPLATE_HEADERS, [])


def _locations_by_reference():
    locations = Location.query.filter(Location.is_active.is_(True)).all()
    refs = {}
    for location in locations:
        refs[str(location.id)] = location
        refs[location.name.strip().lower()] = location
    return refs


def _validate_row(row_number, row, seen_inventory, locations_by_ref):
    data = {key: _clean(value) for key, value in row.items() if key}
    errors = []
    inventory_number = data.get('inventory_number', '')
    if not inventory_number:
        errors.append('Липсва инвентарен №.')
    elif inventory_number in seen_inventory:
        errors.append('Дублиран инвентарен № в импорт файла.')
    else:
        seen_inventory.add(inventory_number)

    if not data.get('name'):
        errors.append('Липсва име/тип на машината.')

    asset_type = data.get('asset_type')
    if asset_type and asset_type not in ASSET_TYPE_OPTIONS:
        errors.append('Невалиден вид актив.')

    if 'purchase_date' in data:
        purchase_date, date_error = _parse_date(data.get('purchase_date'))
        if date_error:
            errors.append(date_error)
        data['purchase_date'] = format_date(purchase_date) if purchase_date else ''

    location_ref = data.get('current_location')
    if location_ref:
        location = locations_by_ref.get(location_ref.lower()) if not location_ref.isdigit() else locations_by_ref.get(location_ref)
        if not location:
            errors.append('Невалидна или неактивна локация.')
        else:
            data['current_location_id'] = location.id
            data['current_location_name'] = location.name
    else:
        data['current_location_id'] = None
        data['current_location_name'] = ''

    existing = Asset.query.filter_by(inventory_number=inventory_number).first() if inventory_number else None
    action = 'update' if existing else 'create'
    return AssetCsvRow(row_number=row_number, action=action, data=data, errors=errors)


def _parse_asset_rows(fieldnames, rows, too_many_rows_message, empty_rows_message):
    preview = AssetCsvPreview()
    if 'inventory_number' not in fieldnames:
        preview.errors.append(f'Липсва задължителна колона: {ASSET_CSV_REQUIRED_COLUMN_LABEL}')
        return preview

    locations_by_ref = _locations_by_reference()
    seen_inventory = set()
    rows_seen = 0
    for rows_seen, row in enumerate(rows, start=1):
        if rows_seen > ASSET_CSV_MAX_ROWS:
            preview.errors.append(too_many_rows_message)
            break
        parsed = _validate_row(rows_seen + 1, row, seen_inventory, locations_by_ref)
        if parsed.errors:
            preview.error_rows.append(parsed)
        else:
            preview.valid_rows.append(parsed)

    if rows_seen == 0:
        preview.errors.append(empty_rows_message)
    return preview


def parse_asset_csv_upload(uploaded_file):
    preview = AssetCsvPreview()
    if not uploaded_file or not getattr(uploaded_file, 'filename', ''):
        preview.errors.append('Изберете CSV или Excel файл.')
        return preview
    if not uploaded_file.filename.lower().endswith('.csv'):
        preview.errors.append('Файлът трябва да бъде CSV.')
        return preview

    uploaded_file.stream.seek(0, io.SEEK_END)
    size = uploaded_file.stream.tell()
    uploaded_file.stream.seek(0)
    if size > ASSET_CSV_MAX_BYTES:
        preview.errors.append('CSV файлът е твърде голям. Максимум 1 MB.')
        return preview

    raw_bytes = uploaded_file.stream.read()
    text, _encoding = _decode_csv_bytes(raw_bytes)
    if not text:
        preview.errors.append('Файлът не може да бъде прочетен. Запазете го като CSV UTF-8 и опитайте отново.')
        return preview

    reader, reader_error = _prepare_csv_reader(text)
    if reader_error:
        preview.errors.append(reader_error)
        return preview

    return _parse_asset_rows(
        reader.fieldnames,
        reader,
        f'CSV файлът съдържа повече от {ASSET_CSV_MAX_ROWS} реда.',
        'CSV файлът няма редове за импорт.',
    )


def parse_asset_xlsx_upload(uploaded_file):
    preview = AssetCsvPreview()
    if not uploaded_file or not getattr(uploaded_file, 'filename', ''):
        preview.errors.append('Изберете CSV или Excel файл.')
        return preview
    if not uploaded_file.filename.lower().endswith('.xlsx'):
        preview.errors.append('Файлът трябва да бъде Excel (.xlsx).')
        return preview

    uploaded_file.stream.seek(0, io.SEEK_END)
    size = uploaded_file.stream.tell()
    uploaded_file.stream.seek(0)
    if size > ASSET_XLSX_MAX_BYTES:
        preview.errors.append('Excel файлът е твърде голям. Максимум 5 MB.')
        return preview

    parsed, parse_error = _prepare_xlsx_rows(uploaded_file.stream.read())
    if parse_error:
        preview.errors.append(parse_error)
        return preview

    return _parse_asset_rows(
        parsed['fieldnames'],
        parsed['rows'],
        f'Excel файлът съдържа повече от {ASSET_CSV_MAX_ROWS} реда.',
        'Excel файлът няма редове за импорт.',
    )


def parse_asset_import_upload(uploaded_file):
    preview = AssetCsvPreview()
    if not uploaded_file or not getattr(uploaded_file, 'filename', ''):
        preview.errors.append('Изберете CSV или Excel файл.')
        return preview

    filename = uploaded_file.filename.lower()
    if filename.endswith('.csv'):
        return parse_asset_csv_upload(uploaded_file)
    if filename.endswith('.xlsx'):
        return parse_asset_xlsx_upload(uploaded_file)

    preview.errors.append('Файлът трябва да бъде CSV или Excel (.xlsx).')
    return preview


def preview_from_payload(payload):
    preview = AssetCsvPreview()
    try:
        rows = json.loads(payload or '[]')
    except json.JSONDecodeError:
        preview.errors.append('Невалидни данни за потвърждение.')
        return preview
    if not isinstance(rows, list):
        preview.errors.append('Невалидни данни за потвърждение.')
        return preview
    if len(rows) > ASSET_CSV_MAX_ROWS:
        preview.errors.append(f'Импортът съдържа повече от {ASSET_CSV_MAX_ROWS} реда.')
        return preview

    locations_by_ref = _locations_by_reference()
    seen_inventory = set()
    for index, row in enumerate(rows, start=2):
        if not isinstance(row, dict):
            preview.error_rows.append(AssetCsvRow(index, 'create', {}, ['Невалиден ред.']))
            continue
        parsed = _validate_row(index, row, seen_inventory, locations_by_ref)
        if parsed.errors:
            preview.error_rows.append(parsed)
        else:
            preview.valid_rows.append(parsed)
    return preview


def apply_asset_csv_import(preview, actor):
    created = 0
    updated = 0
    for row in preview.valid_rows:
        data = row.data
        asset = Asset.query.filter_by(inventory_number=data['inventory_number']).first()
        if asset:
            updated += 1
            is_new = False
        else:
            asset = Asset(
                inventory_number=data['inventory_number'],
                created_by_id=actor.id if actor else None,
            )
            db.session.add(asset)
            created += 1
            is_new = True

        asset.name = data.get('name') or asset.name
        if is_new or 'brand' in data:
            asset.brand = data.get('brand', '')
        if is_new or 'model' in data:
            asset.model = data.get('model', '')
        if 'category' in data:
            asset.category = data.get('category') or None
        elif is_new:
            asset.category = asset.name
        if is_new or 'asset_type' in data:
            asset.asset_type = data.get('asset_type') or asset.asset_type or 'Машина'
        if is_new or 'serial_number' in data:
            asset.serial_number = data.get('serial_number') or None

        if 'purchase_date' in data:
            asset.purchase_date = datetime.strptime(data['purchase_date'], '%Y-%m-%d').date() if data.get('purchase_date') else None

        if 'current_location' in data and data.get('current_location_id'):
            location = db.session.get(Location, int(data['current_location_id']))
            previous_location_id = asset.current_location_id
            asset.current_location_id = location.id
            asset.status = asset_status_from_location(location)
            if previous_location_id != location.id:
                asset.last_moved_at = datetime.utcnow()

        db.session.flush()
        action = 'asset_import_updated' if row.action == 'update' else 'asset_import_created'
        details = f'Импорт на активи от {actor.full_name if actor else "администратор"}.'
        db.session.add(AssetHistory(asset_id=asset.id, action=action, details=details, performed_by_id=actor.id if actor else None))

    db.session.commit()
    return {'created': created, 'updated': updated}
